import os
import time
import xml.etree.ElementTree as ET
from copy import copy, deepcopy
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pyomo.core
from openpyxl.utils.dataframe import dataframe_to_rows

import ExcelReader
import TableDefinition
from CaseStudy import CaseStudy
from TableDefinition import CellStyle, Alignment, Font, Color, Text, Column, NumberFormat, TableDefinition
from printer import Printer

package_directory_ExcelWriter = os.path.dirname(os.path.abspath(__file__))

printer = Printer.getInstance()


class ExcelWriter:
    def __init__(self, excel_definitions_path: str = None):
        """
        Initialize the ExcelWriter with the XML root element.

        :param excel_definitions_path: Path to the TableDefinitions.xml file.
        """
        if excel_definitions_path is None:
            from pathlib import Path
            excel_definitions_path = str(Path(__file__).parent / "TableDefinitions.xml")

            if not Path(excel_definitions_path).exists():
                raise FileNotFoundError(f"TableDefinitions.xml not found at {excel_definitions_path}")

        self.excel_definitions_path = excel_definitions_path
        self.xml_tree = ET.parse(excel_definitions_path)
        self.xml_root = self.xml_tree.getroot()
        self.alignments = Alignment.dict_from_xml(self.xml_root.find("Alignments"))
        self.number_formats = NumberFormat.dict_from_xml(self.xml_root.find("NumberFormats"))
        self.colors = Color.dict_from_xml(self.xml_root.find("Colors"))
        self.fonts = Font.dict_from_xml(self.xml_root.find("Fonts"), self.colors)
        self.texts = Text.dict_from_xml(self.xml_root.find("Texts"))
        self.cell_styles = CellStyle.dict_from_xml(self.xml_root.find("CellStyles"), self.fonts, self.colors, self.number_formats, self.alignments)
        self.columns = Column.dict_from_xml(self.xml_root.find("Columns"), self.cell_styles) | Column.dict_from_xml(self.xml_root.find("PivotColumns"), self.cell_styles)
        self.excel_definitions = TableDefinition.dict_from_xml(self.xml_root.find("TableDefinitions"), self.columns, self.colors, self.cell_styles)
        pass

    @staticmethod
    def __setCellStyle(cell_style: CellStyle, target_cell: openpyxl.cell.cell):
        """
        Set the cell style of a target cell based on the given cell style.

        :param cell_style: CellStyle object containing the style properties.
        :param target_cell: The target cell to which the style will be applied.
        :return: None
        """

        if cell_style.font is not None:
            target_cell.font = openpyxl.styles.fonts.Font(**cell_style.font.__dict__)
        if cell_style.fill is not None:
            target_cell.fill = cell_style.fill
        if cell_style.number_format is not None:
            target_cell.number_format = copy(cell_style.number_format)
        if cell_style.alignment is not None:
            target_cell.alignment = openpyxl.styles.Alignment(**cell_style.alignment.__dict__)

    def _write_Excel_from_definition(self, data: pd.DataFrame, folder_path: str, excel_definition_id: str) -> None:
        """
        Write the given data to an Excel file based on the provided Excel definition.

        :param data: DataFrame containing the data to be written to Excel.
        :param folder_path: Folder path where the Excel file will be saved.
        :param excel_definition_id: ID of the Excel definition to be used.
        :return: None
        """
        start_time = time.time()
        if excel_definition_id not in self.excel_definitions:
            raise ValueError(f"Excel definition '{excel_definition_id}' not found in the definitions. Please define it in the XML file.")
        excel_definition = self.excel_definitions[excel_definition_id]
        wb = openpyxl.Workbook()
        scenarios = data["scenario"].unique()

        data = data.copy()  # Create a copy of the DataFrame to avoid modifying the original data

        # Prepare columns if data should be pivoted
        pivot_columns = []
        target_column = None
        target_column_index = None
        for i, column in enumerate(excel_definition.columns):
            if column.pivoted:
                if target_column is not None:
                    raise ValueError(f"Excel definition '{excel_definition_id}' has (at least) two pivot columns defined: '{target_column.db_name}' and '{column.db_name}'. Only one pivot column is allowed.")
                target_column = column
                target_column_index = i
            else:
                if column.db_name != "NOEXCL":  # Skip first column if it is the (empty and thus unused) placeholder for the excl column
                    pivot_columns.append(column.db_name)

        column_templates = copy(excel_definition.columns)  # Create a copy of the column definitions and adapt this copy for pivoted data
        if target_column is not None:
            data.reset_index(inplace=True)
            data = data.pivot(index=pivot_columns + ["scenario"], columns=target_column.db_name, values="value")
            column_templates.remove(target_column)  # Remove the pivot column from the list of columns
            for i, column in enumerate(data.columns):
                col_definition = copy(target_column)
                col_definition.db_name = column
                col_definition.readable_name = column
                if i != 0:  # Remove description for all but the first pivoted column
                    col_definition.description = None
                column_templates.append(col_definition)  # Add the new column definition to the list of columns

            data.reset_index(inplace=True)

        if len(data) == 0:
            printer.warning(f"No data found for Excel definition '{excel_definition_id}' - writing an empty file.")
            data = pd.DataFrame(columns=[col.db_name for col in column_templates] + ["scenario"])
            scenarios = ["ScenarioA"]

        for scenario_index, scenario in enumerate(scenarios):
            scenario_data = data[data["scenario"] == scenario]

            if scenario_index == 0:
                ws = wb.active
                ws.title = scenario
            else:
                ws = wb.create_sheet(title=scenario)

            # Set sheet properties
            ws.sheet_properties.tabColor = '008080'  # Set tab color
            ws.sheet_view.showGridLines = False  # Hide grid lines
            ws.freeze_panes = "C8"  # Freeze panes at row 8 (below the header)

            # Prepare row heights
            ws.row_dimensions[1].height = 24
            ws.row_dimensions[5].height = excel_definition.description_row_height
            ws.row_dimensions[6].height = 30  # Standard for database behavior row

            # Prepare header columns
            for i, column in enumerate(column_templates):
                if i == 1:  # Column with title text & 'Format' text
                    ws.cell(row=1, column=i + 1, value=excel_definition.sheet_header)
                    ExcelWriter.__setCellStyle(self.cell_styles["title"], ws.cell(row=1, column=i + 1))
                    ws.cell(row=2, column=i + 1, value="Format:")
                    ExcelWriter.__setCellStyle(self.cell_styles["formatDescription"], ws.cell(row=2, column=i + 1))
                elif i == 2:  # Column with format value
                    ExcelWriter.__setCellStyle(self.cell_styles["header"], ws.cell(row=1, column=i + 1))
                    ws.cell(row=2, column=i + 1, value=excel_definition.version)
                    ExcelWriter.__setCellStyle(self.cell_styles["formatValue"], ws.cell(row=2, column=i + 1))
                else:  # Standard header column (no text, just setting the color)
                    ExcelWriter.__setCellStyle(self.cell_styles["header"], ws.cell(row=1, column=i + 1))
                # Set column width
                if column.column_width is not None:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = column.column_width
                # Set notes
                if i == 0 and column.db_name == "excl":
                    ws.cell(row=3, column=i + 1).comment = openpyxl.comments.Comment(self.texts["colHeaderExclDescription"], "")
                if i == 1:
                    ws.cell(row=3, column=i + 1).comment = openpyxl.comments.Comment(self.texts["colHeaderReadableName"], "")
                    ws.cell(row=4, column=i + 1).comment = openpyxl.comments.Comment(self.texts["colHeaderValueSpecifierDB"], "")
                    ws.cell(row=5, column=i + 1).comment = openpyxl.comments.Comment(self.texts["colHeaderDescription"], "")
                    ws.cell(row=6, column=i + 1).comment = openpyxl.comments.Comment(self.texts["colHeaderDBBehavior"], "")
                    ws.cell(row=7, column=i + 1).comment = openpyxl.comments.Comment(self.texts["colHeaderUnit"], "")

                if column.db_name != "NOEXCL":  # Skip first column if it is the (empty and thus unused) placeholder for the excl column
                    # Readable name
                    ws.cell(row=3, column=i + 1, value=column.readable_name)
                    ExcelWriter.__setCellStyle(self.cell_styles["readableName"], ws.cell(row=3, column=i + 1))

                    # Database name
                    ws.cell(row=4, column=i + 1, value=column.db_name)
                    ExcelWriter.__setCellStyle(self.cell_styles["dbName"], ws.cell(row=4, column=i + 1))

                    # Description
                    ws.cell(row=5, column=i + 1, value=column.description)
                    if i != target_column_index:
                        ExcelWriter.__setCellStyle(self.cell_styles["description"], ws.cell(row=5, column=i + 1))
                    else:  # If the column is a pivoted column, set the style without wrapping text
                        cell_style_withou_wrap_text = deepcopy(self.cell_styles["description"])
                        cell_style_withou_wrap_text.alignment.wrap_text = False
                        ExcelWriter.__setCellStyle(cell_style_withou_wrap_text, ws.cell(row=5, column=i + 1))

                    # Database behavior
                    if i != 0:  # Skip db-behavior for the first column (excl)
                        ws.cell(row=6, column=i + 1, value=column.get_db_behavior(self.texts))
                    ExcelWriter.__setCellStyle(self.cell_styles["dbBehavior"], ws.cell(row=6, column=i + 1))

                    # Unit
                    ws.cell(row=7, column=i + 1, value=column.unit)
                    ExcelWriter.__setCellStyle(self.cell_styles["unit"], ws.cell(row=7, column=i + 1))

            # Write data to Excel
            scenario_data = scenario_data.reset_index()
            for i, values in scenario_data.iterrows():
                for j, col in enumerate(column_templates):
                    if col.readable_name is None and j == 0: continue  # Skip first column if it is empty, since it is the (unused) placeholder for the excl column
                    if col.db_name == "excl":  # Excl. column is written by placing 'X' in lines which should be excluded
                        ws.cell(row=i + 8, column=j + 1, value='X' if isinstance(values[col.db_name], str) or not np.isnan(values[col.db_name]) else None)
                    else:
                        ws.cell(row=i + 8, column=j + 1, value=values[col.db_name])
                    ExcelWriter.__setCellStyle(col.cell_style, ws.cell(row=i + 8, column=j + 1))

        path = folder_path + ("/" if not folder_path.endswith("/") else "") + excel_definition.file_name + ".xlsx"
        if not os.path.exists(os.path.dirname(path)) and os.path.dirname(path) != "":
            printer.information(f"Creating folder '{os.path.dirname(path)}'")
            os.makedirs(os.path.dirname(path))  # Create folder if it does not exist
        wb.save(path)
        printer.information(f"Saved Excel file to '{path}' after {time.time() - start_time:.2f} seconds")

    def write_caseStudy(self, cs: CaseStudy, folder_path: str | Path) -> None:
        """
        Write the case study to a folder in LEGO-Excel format.
        :param cs: CaseStudy object containing the data to be written.
        :param folder_path: Path to the folder where the Excel files will be saved.
        :return:
        """
        folder_path = str(folder_path)

        self.write_Global_Scenarios(cs.dGlobal_Scenarios, folder_path)
        self.write_Power_BusInfo(cs.dPower_BusInfo, folder_path)
        self.write_Power_Demand(cs.dPower_Demand, folder_path)
        self.write_Power_Hindex(cs.dPower_Hindex, folder_path)
        if hasattr(cs, "dPower_Inflows"):
            self.write_Power_Inflows(cs.dPower_Inflows, folder_path)
        self.write_Power_Network(cs.dPower_Network, folder_path)
        if hasattr(cs, "dPower_Storage"):
            self.write_Power_Storage(cs.dPower_Storage, folder_path)
        if hasattr(cs, "dPower_ThermalGen"):
            self.write_Power_ThermalGen(cs.dPower_ThermalGen, folder_path)
        if hasattr(cs, "dPower_VRES"):
            self.write_Power_VRES(cs.dPower_VRES, folder_path)
        if hasattr(cs, "dPower_VRESProfiles"):
            self.write_Power_VRESProfiles(cs.dPower_VRESProfiles, folder_path)
        self.write_Power_WeightsK(cs.dPower_WeightsK, folder_path)
        self.write_Power_WeightsRP(cs.dPower_WeightsRP, folder_path)

    def write_Data_Packages(self, dData_Packages: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dData_Packages DataFrame to an Excel file in LEGO format.
        :param dData_Packages: DataFrame containing the dData_Packages data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dData_Packages, folder_path, "Data_Packages")

    def write_Data_Sources(self, dData_Sources: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dData_Sources DataFrame to an Excel file in LEGO format.
        :param dData_Sources: DataFrame containing the dData_Sources data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dData_Sources, folder_path, "Data_Sources")

    def write_Global_Scenarios(self, dGlobal_Scenarios: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dGlobal_Scenarios DataFrame to an Excel file in LEGO format.
        :param dGlobal_Scenarios: DataFrame containing the dGlobal_Scenarios data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dGlobal_Scenarios, folder_path, "Global_Scenarios")

    def write_Power_BusInfo(self, dPower_BusInfo: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_BusInfo DataFrame to an Excel file in LEGO format.
        :param dPower_BusInfo: DataFrame containing the dPower_BusInfo data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_BusInfo, folder_path, "Power_BusInfo")

    def write_Power_Demand(self, dPower_Demand: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_Demand DataFrame to an Excel file in LEGO format.
        :param dPower_Demand: DataFrame containing the dPower_Demand data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """

        self._write_Excel_from_definition(dPower_Demand, folder_path, "Power_Demand")

    def write_Power_Demand_KInRows(self, dPower_Demand_KInRows: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_Demand_KInRows DataFrame to an Excel file in LEGO format.
        :param dPower_Demand_KInRows: DataFrame containing the dPower_Demand_KInRows data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """

        self._write_Excel_from_definition(dPower_Demand_KInRows, folder_path, "Power_Demand_KInRows")

    def write_Power_Hindex(self, dPower_Hindex: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_Hindex DataFrame to an Excel file in LEGO format.
        :param dPower_Hindex: DataFrame containing the dPower_Hindex data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_Hindex, folder_path, "Power_Hindex")

    def write_Power_Inflows(self, dPower_Inflows: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_Inflows DataFrame to an Excel file in LEGO format.
        :param dPower_Inflows: DataFrame containing the dPower_Inflows data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_Inflows, folder_path, "Power_Inflows")

    def write_Power_Inflows_KInRows(self, dPower_Inflows_KInRows: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_Inflows_KInRows DataFrame to an Excel file in LEGO format.
        :param dPower_Inflows_KInRows: DataFrame containing the dPower_Inflows_KInRows data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_Inflows_KInRows, folder_path, "Power_Inflows_KInRows")

    def write_Power_Network(self, dPower_Network: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_Network DataFrame to an Excel file in LEGO format.
        :param dPower_Network: DataFrame containing the dPower_Network data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_Network, folder_path, "Power_Network")

    def write_Power_Storage(self, dPower_Storage: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_Storage DataFrame to an Excel file in LEGO format.
        :param dPower_Storage: DataFrame containing the dPower_Storage data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_Storage, folder_path, "Power_Storage")

    def write_Power_ThermalGen(self, dPower_ThermalGen: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_ThermalGen DataFrame to an Excel file in LEGO format.
        :param dPower_ThermalGen: DataFrame containing the dPower_ThermalGen data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_ThermalGen, folder_path, "Power_ThermalGen")

    def write_Power_VRES(self, dPower_VRES: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_VRES DataFrame to an Excel file in LEGO format.
        :param dPower_VRES: DataFrame containing the dPower_VRES data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_VRES, folder_path, "Power_VRES")

    def write_Power_VRESProfiles(self, dPower_VRESProfiles: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_VRESProfiles DataFrame to an Excel file in LEGO format.
        :param dPower_VRESProfiles: DataFrame containing the dPower_VRESProfiles data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_VRESProfiles, folder_path, "Power_VRESProfiles")

    def write_Power_VRESProfiles_KInRows(self, dPower_VRESProfiles_KInRows: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_VRESProfiles_KInRows DataFrame to an Excel file in LEGO format.
        :param dPower_VRESProfiles_KInRows: DataFrame containing the dPower_VRESProfiles_KInRows data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_VRESProfiles_KInRows, folder_path, "Power_VRESProfiles_KInRows")

    def write_Power_WeightsK(self, dPower_WeightsK: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_WeightsK DataFrame to an Excel file in LEGO format.
        :param dPower_WeightsK: DataFrame containing the dPower_WeightsK data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_WeightsK, folder_path, "Power_WeightsK")

    def write_Power_WeightsRP(self, dPower_WeightsRP: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_WeightsRP DataFrame to an Excel file in LEGO format.
        :param dPower_WeightsRP: DataFrame containing the dPower_WeightsRP data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_WeightsRP, folder_path, "Power_WeightsRP")

    def write_Power_Wind_TechnicalDetails(self, dPower_Wind_TechnicalDetails: pd.DataFrame, folder_path: str) -> None:
        """
        Write the dPower_Wind_TechnicalDetails DataFrame to an Excel file in LEGO format.
        :param dPower_Wind_TechnicalDetails: DataFrame containing the dPower_Wind_TechnicalDetails data.
        :param folder_path: Path to the folder where the Excel file will be saved.
        :return: None
        """
        self._write_Excel_from_definition(dPower_Wind_TechnicalDetails, folder_path, "Power_Wind_TechnicalDetails")

    @staticmethod
    def model_to_excel(model: pyomo.core.Model, target_path: str) -> None:
        """
        Write all variables of the given Pyomo model to an Excel file.

        :param model: The Pyomo model to be written to Excel.
        :param target_path: Path to the target Excel file.
        :return: None
        """
        printer.information(f"Writing model to '{target_path}'")
        wb = openpyxl.Workbook()
        ws = wb.active

        for i, var in enumerate(model.component_objects(pyomo.core.Var, active=True)):
            if i == 0:  # Use the automatically existing sheet for the first variable
                ws.title = str(var)
            else:  # Create a sheet for each (other) variable
                ws = wb.create_sheet(title=str(var))

            # Prepare the data from the model
            data = [(j, v.value if not v.stale else None) for j, v in var.items()]

            # Extract parameter names from the variable's index structure
            param_names = []

            if var.is_indexed():
                index_set = var.index_set()

                try:
                    # Get names from the index set
                    if hasattr(index_set, 'subsets') and index_set.subsets():
                        for idx, subset in enumerate(index_set.subsets()):
                            if subset.domain.dimen is not None:
                                for i, domain in enumerate(subset.domain.subsets()):
                                    param_names.append(f"{subset.name}[{i}]: {domain.name}")
                            else:
                                param_names.append(subset.name)
                        param_names.append(str(var))
                except (AttributeError, TypeError):
                    if len(data) > 0:
                        # Determine from actual data structure
                        col_number = len(data[0][0]) if not isinstance(data[0][0], str) else 1
                        param_names = [f"index_{j}" for j in range(col_number)] + [str(var)]
                    else:
                        param_names = []

            # Create header row with parameter names
            ws.append(param_names)

            # Handle data writing
            if len(data) == 0:
                # Create a row showing "No entries" for each parameter
                ws.append(["No entries"] * len(param_names))
            else:
                # Write data to the sheet
                for j, v in data:
                    ws.append(([j_index for j_index in j] if not isinstance(j, str) else [j]) + [v])

        wb.save(target_path)


if __name__ == "__main__":
    import argparse
    from rich_argparse import RichHelpFormatter

    parser = argparse.ArgumentParser(description="Re-write all files in given folder and compare against source", formatter_class=RichHelpFormatter)
    parser.add_argument("caseStudyFolder", type=str, help="Path to folder containing data for LEGO model.")
    parser.add_argument("excelDefinitionsPath", type=str, help="Path to the Excel definitions XML file. Uses default if none is supplied.", nargs="?")
    parser.add_argument("--dontCheckFormatting", action="store_true", help="Do not check formatting of the Excel files. Only check if they are equal.")
    parser.add_argument("--dontFailOnWrongVersion", action="store_true", help="Do not fail if the version in the Excel file does not match the version in the XML definitions file.")
    parser.add_argument("--precision", type=float, default=1e-6, help="Precision for comparing floating point values, default is 1e-6")
    args = parser.parse_args()

    printer.set_width(300)

    if not args.caseStudyFolder.endswith("/"):
        args.caseStudyFolder += "/"
    printer.information(f"Loading case study from '{args.caseStudyFolder}'")

    if args.excelDefinitionsPath is None:
        ew = ExcelWriter()
    else:
        ew = ExcelWriter(args.excelDefinitionsPath)
        printer.information(f"Loading Excel definitions from '{args.excelDefinitionsPath}'")
    printer.separator()

    combinations = [
        ("Data_Packages", f"{args.caseStudyFolder}Data_Packages.xlsx", ExcelReader.get_Data_Packages, ew.write_Data_Packages),
        ("Data_Sources", f"{args.caseStudyFolder}Data_Sources.xlsx", ExcelReader.get_Data_Sources, ew.write_Data_Sources),
        ("Global_Scenarios", f"{args.caseStudyFolder}Global_Scenarios.xlsx", ExcelReader.get_Global_Scenarios, ew.write_Global_Scenarios),
        ("Power_BusInfo", f"{args.caseStudyFolder}Power_BusInfo.xlsx", ExcelReader.get_Power_BusInfo, ew.write_Power_BusInfo),
        ("Power_Demand", f"{args.caseStudyFolder}Power_Demand.xlsx", ExcelReader.get_Power_Demand, ew.write_Power_Demand),
        ("Power_Demand_KInRows", f"{args.caseStudyFolder}Power_Demand_KInRows.xlsx", ExcelReader.get_Power_Demand_KInRows, ew.write_Power_Demand_KInRows),
        ("Power_Hindex", f"{args.caseStudyFolder}Power_Hindex.xlsx", ExcelReader.get_Power_Hindex, ew.write_Power_Hindex),
        ("Power_Inflows", f"{args.caseStudyFolder}Power_Inflows.xlsx", ExcelReader.get_Power_Inflows, ew.write_Power_Inflows),
        ("Power_Inflows_KInRows", f"{args.caseStudyFolder}Power_Inflows_KInRows.xlsx", ExcelReader.get_Power_Inflows_KInRows, ew.write_Power_Inflows_KInRows),
        ("Power_Network", f"{args.caseStudyFolder}Power_Network.xlsx", ExcelReader.get_Power_Network, ew.write_Power_Network),
        ("Power_Storage", f"{args.caseStudyFolder}Power_Storage.xlsx", ExcelReader.get_Power_Storage, ew.write_Power_Storage),
        ("Power_ThermalGen", f"{args.caseStudyFolder}Power_ThermalGen.xlsx", ExcelReader.get_Power_ThermalGen, ew.write_Power_ThermalGen),
        ("Power_VRES", f"{args.caseStudyFolder}Power_VRES.xlsx", ExcelReader.get_Power_VRES, ew.write_Power_VRES),
        ("Power_VRESProfiles", f"{args.caseStudyFolder}Power_VRESProfiles.xlsx", ExcelReader.get_Power_VRESProfiles, ew.write_Power_VRESProfiles),
        ("Power_VRESProfiles_KInRows", f"{args.caseStudyFolder}Power_VRESProfiles_KInRows.xlsx", ExcelReader.get_Power_VRESProfiles_KInRows, ew.write_Power_VRESProfiles_KInRows),
        ("Power_WeightsK", f"{args.caseStudyFolder}Power_WeightsK.xlsx", ExcelReader.get_Power_WeightsK, ew.write_Power_WeightsK),
        ("Power_WeightsRP", f"{args.caseStudyFolder}Power_WeightsRP.xlsx", ExcelReader.get_Power_WeightsRP, ew.write_Power_WeightsRP),
        ("Power_Wind_TechnicalDetails", f"{args.caseStudyFolder}Power_Wind_TechnicalDetails.xlsx", ExcelReader.get_Power_Wind_TechnicalDetails, ew.write_Power_Wind_TechnicalDetails)
    ]

    for excel_definition_id, file_path, read, write in combinations:
        printer.information(f"Writing '{excel_definition_id}', read from '{file_path}'")
        data = read(file_path, True, not args.dontFailOnWrongVersion)
        write(data, f"{args.caseStudyFolder}output")

        printer.information(f"Comparing '{args.caseStudyFolder}output/{excel_definition_id}.xlsx' against source file '{file_path}'")
        if not os.path.exists(file_path):
            printer.warning(f"Input file '{file_path}' does not exist - skipping comparison")
        else:
            filesEqual = ExcelReader.compare_Excels(file_path, f"{args.caseStudyFolder}output/{excel_definition_id}.xlsx", args.dontCheckFormatting, args.precision)
            if filesEqual:
                printer.success(f"Excel files are equal")
            else:
                printer.error(f"Excel files are NOT equal - see above for details")

        printer.separator()
