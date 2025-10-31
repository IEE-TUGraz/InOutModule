import time

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from printer import Printer

printer = Printer.getInstance()


def check_LEGOExcel_version(xls: pd.ExcelFile, sheet_name: str, version_specifier: str, excel_file_path: str, fail_on_wrong_version: bool = False):
    """
    Check if a specific sheet in an open Excel file has the correct version specifier.
    :param xls: The open pandas.ExcelFile object
    :param sheet_name: The name of the sheet to check
    :param version_specifier: Expected version specifier (e.g., "v0.1.0")
    :param excel_file_path: Path to the Excel file (for error logging)
    :param fail_on_wrong_version: If True, raise an error if the version does not match
    """
    try:
        # Read only cell C2 (row=2, col=3) from the specified sheet
        version_cell = pd.read_excel(xls, sheet_name=sheet_name, usecols="C", skiprows=1, nrows=1, header=None).iloc[0, 0]
    except Exception as e:
        printer.error(f"Could not read version cell [C2] from sheet '{sheet_name}' in '{excel_file_path}'. Error: {e}")
        version_cell = None

    if version_cell != version_specifier:
        if fail_on_wrong_version:
            raise ValueError(
                f"Excel file '{excel_file_path}' (sheet '{sheet_name}') does not have the correct version specifier. "
                f"Expected '{version_specifier}' but got '{version_cell}'.")
        else:
            printer.error(
                f"Excel file '{excel_file_path}' (sheet '{sheet_name}') does not have the correct version specifier. "
                f"Expected '{version_specifier}' but got '{version_cell}'.")
            printer.error(f"Trying to work with it any way, but this can have unintended consequences!")


def __read_non_pivoted_file(excel_file_path: str, version_specifier: str, indices: list[str], has_excl_column: bool,
                            keep_excl_columns: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read a non-pivoted Excel file and return the data as a DataFrame.
    :param excel_file_path: Path to the Excel file
    :param version_specifier: Version specifier to check against the Excel file
    :param indices: List of columns to set as index in the DataFrame
    :param has_excl_column: If True, the DataFrame has an "Excl." column that indicates whether a row should be excluded
    :param keep_excl_columns: If True, keep the "Excl." column in the DataFrame
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: DataFrame containing the data from the Excel file
    """
    xls = pd.ExcelFile(excel_file_path, engine="calamine")
    data = pd.DataFrame()

    for scenario in xls.sheet_names:  # Iterate through all sheets, i.e., through all scenarios
        if scenario.startswith("~"):
            printer.warning(f"Skipping sheet '{scenario}' from '{excel_file_path}' because it starts with '~'.")
            continue

        check_LEGOExcel_version(xls, scenario, version_specifier, excel_file_path, fail_on_wrong_version)

        df = pd.read_excel(xls, skiprows=[0, 1, 2, 4, 5, 6], sheet_name=scenario)
        if has_excl_column:
            if not keep_excl_columns:
                df = df[df["excl"].isnull()]  # Only keep rows that are not excluded (i.e., have no value in the "Excl." column)
        else:
            df = df.drop(df.columns[0], axis=1)  # Drop the first column (which is empty)
        df = df.set_index(indices) if len(indices) > 0 else df
        df["scenario"] = scenario

        data = pd.concat([data, df], ignore_index=False)  # Append the DataFrame to the main DataFrame

    return data


def __read_pivoted_file(excel_file_path: str, version_specifier: str, indices: list[str], pivoted_variable_name: str, melt_indices: list[str], has_excl_column: bool,
                        keep_excluded_columns: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read a pivoted Excel file and return the data as a DataFrame.
    :param excel_file_path: Path to the Excel file
    :param version_specifier: Version specifier to check against the Excel file
    :param indices: List of columns to set as index in the DataFrame
    :param pivoted_variable_name: Name of the variable that was pivoted in the Excel
    :param melt_indices: List of columns to keep as identifiers when melting the DataFrame
    :param has_excl_column: If True, the DataFrame has an "Excl." column that indicates whether a row should be excluded
    :param keep_excluded_columns: If True, keep the "Excl." column in the DataFrame
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: DataFrame containing the data from the Excel file
    """
    df = __read_non_pivoted_file(excel_file_path, version_specifier, [], has_excl_column, keep_excluded_columns, fail_on_wrong_version)

    df = df.melt(id_vars=melt_indices + ["scenario"], var_name=pivoted_variable_name, value_name="value")
    df = df.set_index(indices)
    return df


def get_Data_Packages(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dData_Packages data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dData_Packages
    """
    dData_Packages = __read_non_pivoted_file(excel_file_path, "v0.1.0", ["dataPackage"], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Data_Packages', although nothing is excluded anyway - please check if this is intended.")

    return dData_Packages


def get_Data_Sources(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dData_Sources data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dData_Sources
    """
    dData_Sources = __read_non_pivoted_file(excel_file_path, "v0.2.0", ["dataSource"], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Data_Sources', although nothing is excluded anyway - please check if this is intended.")

    return dData_Sources


def get_Global_Scenarios(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dGlobal_Scenarios data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Do not exclude any entries which are marked to be excluded in the Excel file
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dGlobal_Scenarios
    """
    dGlobal_Scenarios = __read_non_pivoted_file(excel_file_path, "v0.1.0", ["scenarioID"], True, keep_excluded_entries, fail_on_wrong_version)

    # Check that there is only one sheet with the name 'Scenario'
    check = dGlobal_Scenarios["scenario"].to_numpy()
    if not (check[0] == check).all():
        raise ValueError(f"There are multiple or falsely named sheets for '{excel_file_path}'. There should only be one sheet with the name 'Scenario', please check the Excel file.")

    return dGlobal_Scenarios


def get_Power_BusInfo(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_BusInfo data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Do not exclude any entries which are marked to be excluded in the Excel file
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_BusInfo
    """
    dPower_BusInfo = __read_non_pivoted_file(excel_file_path, "v0.1.2", ["i"], True, keep_excluded_entries, fail_on_wrong_version)

    return dPower_BusInfo


def get_Power_Demand(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_Demand data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_Demand
    """
    dPower_Demand = __read_pivoted_file(excel_file_path, "v0.1.4", ['rp', 'k', 'i'], 'k', ['rp', 'i', 'dataPackage', 'dataSource', 'id'], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Power_Demand', although nothing is excluded anyway - please check if this is intended.")

    return dPower_Demand


def get_Power_Demand_KInRows(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_Demand_KInRows data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_Demand_KInRows
    """
    dPower_Demand_KInRows = __read_pivoted_file(excel_file_path, "v0.1.4", ['rp', 'k', 'i'], 'i', ['rp', 'k', 'dataPackage', 'dataSource', 'id'], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Power_Demand_KInRows', although nothing is excluded anyway - please check if this is intended.")

    return dPower_Demand_KInRows


def get_Power_Hindex(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_Hindex data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_Hindex
    """
    dPower_Hindex = __read_non_pivoted_file(excel_file_path, "v0.1.3", ["p", "rp", "k"], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Power_Hindex', although nothing is excluded anyway - please check if this is intended.")

    return dPower_Hindex


def get_Power_ImportExport(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_ImportExport data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_ImportExport
    """
    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Power_ImportExport', although nothing is excluded anyway - please check if this is intended.")

    version_specifier = "v0.0.1"
    xls = pd.ExcelFile(excel_file_path, engine="calamine")
    data = pd.DataFrame()

    for scenario in xls.sheet_names:  # Iterate through all sheets, i.e., through all scenarios
        if scenario.startswith("~"):
            printer.warning(f"Skipping sheet '{scenario}' from '{excel_file_path}' because it starts with '~'.")
            continue

        check_LEGOExcel_version(xls, scenario, version_specifier, excel_file_path, fail_on_wrong_version)

        # Read row 3 (information about hubs and nodes)
        hub_i_df = pd.read_excel(xls, skiprows=[0, 1, 3], nrows=2, sheet_name=scenario)
        hub_i = []
        hubs = []
        i = 6  # Start checking from column 7 (index 6, zero-based)
        while i < hub_i_df.shape[1]:
            hubs.append(hub_i_df.columns[i])
            hub_i.append((hub_i_df.columns[i], hub_i_df.columns[i + 1]))
            if "Unnamed" not in hub_i_df.columns[i + 2]:
                raise ValueError(f"Power_ImportExport: Expected pairs of columns for hub and i, but found an unexpected text '{hub_i_df.columns[i + 2]}' at column index {get_column_letter(i + 3)}. Please check the Excel file format.")
            i += 3  # Move to the next pair (skip the "Unnamed" column)

        if len(hubs) != len(set(hubs)):
            raise ValueError(f"Power_ImportExport: Found duplicate hub names in the header row. Hubs must be unique. Please check the Excel file.")

        df = pd.read_excel(xls, skiprows=[0, 1, 2, 4, 5, 6], sheet_name=scenario)
        df = df.drop(df.columns[0], axis=1)  # Drop the first column (which is empty)

        for i, col in enumerate(df.columns):
            if i < 5:
                continue  # Skip the first five columns
            hub = hub_i[(i - 5) // 3][0]
            node = hub_i[(i - 5) // 3][1]

            match (i - 5) % 3:
                case 0:
                    if "ImpExpMinimum" not in col:
                        raise ValueError(f"Power_ImportExport: Expected column 'ImpExpMinimum' at column index {get_column_letter(i + 2)}, but found '{col}'. Please check the Excel file format.")
                    col_name = "ImpExpMinimum"
                case 1:
                    if "ImpExpMaximum" not in col:
                        raise ValueError(f"Power_ImportExport: Expected column 'ImpExpMaximum' at column index {get_column_letter(i + 2)}, but found '{col}'. Please check the Excel file format.")
                    col_name = "ImpExpMaximum"
                case 2:
                    if "ImpExpPrice" not in col:
                        raise ValueError(f"Power_ImportExport: Expected column 'ImpExpPrice' at column index {get_column_letter(i + 2)}, but found '{col}'. Please check the Excel file format.")
                    col_name = "ImpExpPrice"
                case _:
                    raise ValueError("This should never happen.")

            if "@" in hub:
                raise ValueError(f"Power_ImportExport: Found '@' in hub name {hub}, which is not allowed. Please rename it.")
            elif "@" in node:
                raise ValueError(f"Power_ImportExport: Found '@' in node name {node}, which is not allowed. Please rename it.")
            df = df.rename(columns={col: f"{hub}@{node}@{col_name}"})

        df = df.melt(id_vars=["id", "rp", "k", "dataPackage", "dataSource"])

        df[["hub", "i", "valueType"]] = df["variable"].str.split("@", expand=True)  # Split the variable column into hub, i and valueType

        df = df.pivot(index=["id", "rp", "k", "dataPackage", "dataSource", "hub", "i"], columns="valueType", values="value")
        df.columns.name = None  # Fix name of columns/indices (which are altered through pivot)

        df["scenario"] = scenario

        df = df.reset_index().set_index(["hub", "i", "rp", "k"])  # Set multiindex

        data = pd.concat([data, df], ignore_index=False)  # Append the DataFrame to the main DataFrame

    return data


def get_Power_Inflows(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_Inflows data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_Inflows
    """
    dPower_Inflows = __read_pivoted_file(excel_file_path, "v0.1.0", ['rp', 'k', 'g'], 'k', ['rp', 'g', 'dataPackage', 'dataSource', 'id'], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Power_Inflows', although nothing is excluded anyway - please check if this is intended.")

    return dPower_Inflows


def get_Power_Inflows_KInRows(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_Inflows_KInRows data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_Inflows
    """
    dPower_Inflows_KInRows = __read_pivoted_file(excel_file_path, "v0.1.0", ['rp', 'k', 'g'], 'g', ['rp', 'k', 'dataPackage', 'dataSource', 'id'], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Power_Inflows_KInRows', although nothing is excluded anyway - please check if this is intended.")

    return dPower_Inflows_KInRows


def get_Power_Network(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_Network data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Do not exclude any entries which are marked to be excluded in the Excel file
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_Network
    """
    dPower_Network = __read_non_pivoted_file(excel_file_path, "v0.1.2", ["i", "j", "c"], True, keep_excluded_entries, fail_on_wrong_version)

    return dPower_Network


def get_Power_Storage(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_Storage data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Do not exclude any entries which are marked to be excluded in the Excel file
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_Storage
    """
    dPower_Storage = __read_non_pivoted_file(excel_file_path, "v0.0.2", ["g"], True, keep_excluded_entries, fail_on_wrong_version)

    return dPower_Storage


def get_Power_ThermalGen(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_ThermalGen data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Do not exclude any entries which are marked to be excluded in the Excel file
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_thermalGen
    """
    dPower_ThermalGen = __read_non_pivoted_file(excel_file_path, "v0.1.1", ["g"], True, keep_excluded_entries, fail_on_wrong_version)

    return dPower_ThermalGen


def get_Power_VRES(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_VRES data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Do not exclude any entries which are marked to be excluded in the Excel file
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_VRES
    """
    dPower_VRES = __read_non_pivoted_file(excel_file_path, "v0.1.0", ["g"], True, keep_excluded_entries, fail_on_wrong_version)

    return dPower_VRES


def get_Power_VRESProfiles(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_VRESProfiles data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_VRESProfiles
    """
    dPower_VRESProfiles = __read_pivoted_file(excel_file_path, "v0.1.1", ['rp', 'k', 'g'], 'k', ['rp', 'g', 'dataPackage', 'dataSource', 'id'], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Power_VRESProfiles', although nothing is excluded anyway - please check if this is intended.")

    return dPower_VRESProfiles


def get_Power_VRESProfiles_KInRows(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_VRESProfiles_KInRows data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_VRESProfiles_KInRows
    """
    dPower_VRESProfiles_KInRows = __read_pivoted_file(excel_file_path, "v0.1.1", ['rp', 'k', 'g'], 'g', ['rp', 'k', 'dataPackage', 'dataSource', 'id'], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Power_VRESProfiles_KInRows', although nothing is excluded anyway - please check if this is intended.")

    return dPower_VRESProfiles_KInRows


def get_Power_WeightsK(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_WeightsK data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_WeightsK
    """
    dPower_WeightsK = __read_non_pivoted_file(excel_file_path, "v0.1.4", ["k"], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Power_WeightsK', although nothing is excluded anyway - please check if this is intended.")

    return dPower_WeightsK


def get_Power_WeightsRP(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_WeightsRP data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Unused but kept for compatibility with other functions
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_WeightsRP
    """
    dPower_WeightsRP = __read_non_pivoted_file(excel_file_path, "v0.1.3", ["rp"], False, False, fail_on_wrong_version)

    if keep_excluded_entries:
        printer.warning("'keep_excluded_entries' is set for 'get_Power_WeightsRP', although nothing is excluded anyway - please check if this is intended.")

    return dPower_WeightsRP


def get_Power_Wind_TechnicalDetails(excel_file_path: str, keep_excluded_entries: bool = False, fail_on_wrong_version: bool = False) -> pd.DataFrame:
    """
    Read the dPower_Wind_TechnicalDetails data from the Excel file.
    :param excel_file_path: Path to the Excel file
    :param keep_excluded_entries: Do not exclude any entries which are marked to be excluded in the Excel file
    :param fail_on_wrong_version: If True, raise an error if the version of the Excel file does not match the expected version
    :return: dPower_Wind_TechnicalDetails
    """
    dPower_Wind_TechnicalDetails = __read_non_pivoted_file(excel_file_path, "v0.1.0", ["g"], True, keep_excluded_entries, fail_on_wrong_version)

    return dPower_Wind_TechnicalDetails


def compare_Excels(source_path: str, target_path: str, dont_check_formatting: bool = False, precision: float = 1e-6) -> bool:
    """
    Compare two Excel files for differences in formatting and values.
    :param source_path: Path to the source Excel file
    :param target_path: Path to the target Excel file
    :param dont_check_formatting: If True, skip formatting checks
    :param precision: Precision for floating point comparison
    :return: True if the files are equal, False otherwise
    """
    start_time = time.time()
    source = load_workbook(source_path)
    target = load_workbook(target_path)

    equal = True

    for sheet in source.sheetnames:
        source_sheet = source[sheet]
        if sheet not in target.sheetnames:
            printer.error(f"Sheet '{sheet}' not found in target file '{target_path}'")
            equal = False
            continue
        target_sheet = target[sheet]

        for row in range(1, min(source_sheet.max_row, target_sheet.max_row) + 1):
            if not dont_check_formatting:
                if source_sheet.row_dimensions[row].height != target_sheet.row_dimensions[row].height:
                    printer.error(f"Mismatch in row height at {sheet}/row {row}: {source_sheet.row_dimensions[row].height} != {target_sheet.row_dimensions[row].height}")
                    equal = False

            for col in range(1, min(source_sheet.max_column, target_sheet.max_column) + 1):
                source_cell = source_sheet.cell(row=row, column=col)
                target_cell = target_sheet.cell(row=row, column=col)

                # Value
                if source_cell.value != target_cell.value:
                    if (isinstance(source_cell.value, float) or isinstance(source_cell.value, int)) and (isinstance(target_cell.value, float) or isinstance(target_cell.value, int)):
                        if abs(source_cell.value - target_cell.value) / (source_cell.value if source_cell.value != 0 else 1) >= precision:
                            source_value = str(source_cell.value).replace("[", r"\[")  # Required to prevent rich from interpreting brackets as style definitions
                            target_value = str(target_cell.value).replace("[", r"\[")
                            printer.error(f"Mismatch in value at {sheet}/{source_cell.coordinate}: {source_value} != {target_value}")
                            equal = False
                    else:
                        source_value = str(source_cell.value).replace("[", r"\[")  # Required to prevent rich from interpreting brackets as style definitions
                        target_value = str(target_cell.value).replace("[", r"\[")
                        printer.error(f"Mismatch in value at {sheet}/{source_cell.coordinate}: {source_value} != {target_value}")
                        equal = False

                if not dont_check_formatting:
                    # Font
                    for k, v in source_cell.font.__dict__.items():
                        if k == "color" and v is not None:
                            for k2, v2 in v.__dict__.items():
                                if v2 != getattr(target_cell.font.color, k2):
                                    printer.error(f"Mismatch in font color at {sheet}/{source_cell.coordinate}: {v2} != {getattr(target_cell.font.color, k2)}")
                                    equal = False
                        elif getattr(target_cell.font, k) != v:
                            printer.error(f"Mismatch in font property '{k}' at {sheet}/{source_cell.coordinate}: {getattr(target_cell.font, k)} != {v}")
                            equal = False

                    # Fill
                    for k, v in source_cell.fill.__dict__.items():
                        if k == "color" and v is not None:
                            for k2, v2 in v.__dict__.items():
                                if v2 != getattr(target_cell.fill.color, k2):
                                    printer.error(f"Mismatch in fill color at {sheet}/{source_cell.coordinate}: {v2} != {getattr(target_cell.fill.color, k2)}")
                                    equal = False
                        elif getattr(target_cell.fill, k) != v:
                            printer.error(f"Mismatch in fill property '{k}' at {sheet}/{source_cell.coordinate}: {getattr(target_cell.fill, k)} != {v}")
                            equal = False

                    # Number format
                    if source_cell.number_format != target_cell.number_format:
                        printer.error(f"Mismatch in number format at {sheet}/{source_cell.coordinate}: {source_cell.number_format} != {target_cell.number_format}")
                        equal = False

                    # Alignment
                    for k, v in source_cell.alignment.__dict__.items():
                        if getattr(target_cell.alignment, k) != v:
                            printer.error(f"Mismatch in alignment property '{k}' at {sheet}/{source_cell.coordinate}: {getattr(target_cell.alignment, k)} != {v}")
                            equal = False

                    # Comment
                    if ((source_cell.comment is None and target_cell.comment is not None) or
                            (source_cell.comment is not None and target_cell.comment is None) or
                            (source_cell.comment != target_cell.comment)):
                        printer.error(f"Mismatch in comment at {sheet}/{source_cell.coordinate}: {source_cell.comment} != {target_cell.comment}")
                        equal = False

                    # Column width
                    if row == 1:  # Only need to check column width for the first row
                        source_columnwidth = source_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width
                        for group in source_sheet.column_groups:
                            start, end = group.split(":")
                            start = openpyxl.utils.column_index_from_string(start)
                            end = openpyxl.utils.column_index_from_string(end)
                            if start < col <= end:
                                source_columnwidth = source_sheet.column_dimensions[openpyxl.utils.get_column_letter(start)].width
                                break

                        target_columnwidth = target_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width
                        for group in target_sheet.column_groups:
                            start, end = group.split(":")
                            start = openpyxl.utils.column_index_from_string(start)
                            end = openpyxl.utils.column_index_from_string(end)
                            if start < col <= end:
                                target_columnwidth = target_sheet.column_dimensions[openpyxl.utils.get_column_letter(start)].width
                                break
                        if source_columnwidth != target_columnwidth:
                            printer.error(f"Mismatch in column width at {sheet}/column {col}: {source_columnwidth} != {target_columnwidth}")
                            equal = False
        if source_sheet.max_column != target_sheet.max_column:
            printer.error(f"Target sheet '{sheet}' has {abs(source_sheet.max_column - target_sheet.max_column)} {"more" if source_sheet.max_column > target_sheet.max_column else "less"} columns ({target_sheet.max_column} in total) than source sheet ({source_sheet.max_column} in total)")
            equal = False
        if source_sheet.max_row != target_sheet.max_row:
            printer.error(f"Target sheet '{sheet}' has {abs(source_sheet.max_row - target_sheet.max_row)} {"more" if source_sheet.max_row > target_sheet.max_row else "less"} rows ({target_sheet.max_row} in total) than source sheet ({source_sheet.max_row} in total)")
            equal = False

    printer.information(f"Compared Excel file '{source_path}' to '{target_path}' in {time.time() - start_time:.2f} seconds")
    return equal
