import xml
from copy import copy
from typing import Optional, Self

import openpyxl


class Alignment(openpyxl.styles.Alignment):
    """
    Wrapper for openpyxl.styles.Alignment to provide a more user-friendly interface.
    """

    @classmethod
    def dict_from_xml(cls, alignments: xml.etree.ElementTree.Element) -> dict[str, Self]:
        """
        Converts a list of XML alignment elements into a dictionary of Alignment objects.
        :param alignments: XML element containing alignment definitions.
        :return: A dictionary mapping alignment IDs to Alignment objects.
        """
        return {alignment.get("id"): Alignment(horizontal=alignment.find("horizontal").text if alignment.find("horizontal").text is not None else None,
                                               vertical=alignment.find("vertical").text if alignment.find("vertical").text is not None else None,
                                               wrap_text=alignment.find("wrapText").text == "True" if alignment.find("wrapText").text is not None else None,
                                               indent=float(alignment.find("indent").text) if alignment.find("indent").text is not None else 0) for alignment in alignments
                }


class NumberFormat(str):
    """
    Wrapper for number format strings to provide a more user-friendly interface.
    """

    @classmethod
    def dict_from_xml(cls, number_formats: xml.etree.ElementTree.Element) -> dict[str, Self]:
        """
        Converts a list of XML number format elements into a dictionary of NumberFormat objects.
        :param number_formats: XML element containing number format definitions.
        :return: A dictionary mapping number format IDs to NumberFormat objects.
        """
        return {number_format.get("id"): NumberFormat(number_format.text) for number_format in number_formats}


class Color(str):
    """
    Wrapper for hex-formatted color strings to provide a more user-friendly interface.
    """

    def to_patternFill(self) -> openpyxl.styles.fills.PatternFill:
        """
        Converts the hex color string to an openpyxl PatternFill object.
        :return: An openpyxl PatternFill object with the specified color.
        """
        return openpyxl.styles.PatternFill(start_color=self, end_color=self, fill_type="solid")

    @classmethod
    def dict_from_xml(cls, colors: xml.etree.ElementTree.Element) -> dict[str, Self]:
        """
        Converts a list of XML color elements into a dictionary of Color objects.
        :param colors: XML element containing color definitions.
        :return: A dictionary mapping color IDs to Color objects.
        """
        return {color.get("id"): Color(color.text) for color in colors}


class Font(openpyxl.styles.fonts.Font):
    """
    Wrapper for openpyxl.styles.fonts.Font to provide a more user-friendly interface.
    """

    @classmethod
    def dict_from_xml(cls, fonts: xml.etree.ElementTree.Element, color_dict: dict[str, Color]) -> dict[str, Self]:
        """
        Converts a list of XML font elements into a dictionary of Font objects.
        :param fonts: XML element containing font definitions.
        :param color_dict: Dictionary mapping color IDs to Color objects.
        :return: A dictionary mapping font IDs to Font objects.
        """
        return {font.get("id"): Font(name=font.find("name").text,
                                     size=int(font.find("size").text),
                                     bold=font.find("bold").text == "True",
                                     italic=font.find("italic").text == "True",
                                     color=color_dict[font.find("Color").text] if font.find("Color").text is not None else None) for font in fonts}


class Text(str):
    """
    Wrapper for text strings to provide a more user-friendly interface.
    """

    @classmethod
    def dict_from_xml(cls, texts: xml.etree.ElementTree.Element) -> dict[str, Self]:
        """
        Converts a list of XML text elements into a dictionary of Text objects.
        :param texts: XML element containing text definitions.
        :return: A dictionary mapping text IDs to Text objects.
        """
        return {text.get("id"): Text(text.text) for text in texts}


class CellStyle:
    def __init__(self, font: openpyxl.styles.fonts.Font,
                 fill: openpyxl.styles.fills.PatternFill,
                 number_format: Optional[str],
                 alignment: openpyxl.styles.alignment.Alignment):
        self.font = font
        self.fill = fill
        self.number_format = number_format
        self.alignment = alignment

    def get_copy_with_scenario_dependent(self, scenario_dependent: bool, color_dict: dict[str, Color]) -> Self:
        """
        Sets the fill color of the cell style based on whether the cell is scenario dependent or not and returns a copy of the object.

        :param scenario_dependent: Boolean indicating if the cell is scenario dependent.
        :param color_dict: Dictionary mapping color IDs to Color objects.
        :return: Copy of the updated CellStyle object.
        """
        obj = copy(self)
        if scenario_dependent:
            obj.fill = color_dict["darkBlue"].to_patternFill()
        else:
            obj.fill = color_dict["lightGreen"].to_patternFill()
        return obj

    @classmethod
    def dict_from_xml(cls, cell_styles: xml.etree.ElementTree.Element, font_dict: Optional[dict[str, Font]], color_dict: Optional[dict[str, Color]], number_format_dict: Optional[dict[str, str]], alignment_dict: Optional[dict[str, Alignment]]) -> dict[str, Self]:
        """
        Converts a list of XML cell style elements into a dictionary of CellStyle objects.
        :param cell_styles: XML element containing cell style definitions.
        :return: A dictionary mapping cell style IDs to CellStyle objects.
        """
        return {cell_style.get("id"): CellStyle(font=font_dict[cell_style.find("Font").text] if cell_style.find("Font").text is not None else None,
                                                fill=color_dict[cell_style.find("Color").text].to_patternFill() if cell_style.find("Color").text is not None else None,
                                                number_format=number_format_dict[cell_style.find("NumberFormat").text] if cell_style.find("NumberFormat").text is not None else None,
                                                alignment=alignment_dict[cell_style.find("Alignment").text] if cell_style.find("Alignment").text is not None else None) for cell_style in cell_styles}


class Column:
    def __init__(self, readable_name: str, db_name: str, description: str, unit: str, column_width: float, cell_style: CellStyle, scenario_dependent: bool = False):
        self.readable_name = readable_name
        self.db_name = db_name
        self.description = description
        self.unit = unit
        self.column_width = column_width + 0.7109375  # Difference between Excel's default font and the shown column width (see https://foss.heptapod.net/openpyxl/openpyxl/-/issues/293)
        self.cell_style = cell_style
        self.scenario_dependent = scenario_dependent

    def get_copy_with_scenario_dependent(self, scenario_dependent: bool, color_dict: dict[str, Color]) -> Self:
        """
        Sets the scenario_dependent attribute of the column and returns a copy of the object. Important since the column is used in multiple ExcelDefinition objects.
        :param scenario_dependent: Boolean indicating if the column is scenario dependent.
        :param color_dict: Dictionary mapping color IDs to Color objects.
        :return: Copy of the updated Column object.
        """
        obj = copy(self)
        obj.scenario_dependent = scenario_dependent
        if obj.cell_style is not None and obj.db_name != "id":  # Don't set the color for the ID column
            obj.cell_style = obj.cell_style.get_copy_with_scenario_dependent(scenario_dependent, color_dict)
        return obj

    def get_db_behavior(self, text_dict: dict[str, Text]) -> str:
        """
        Returns the database behavior of the column.
        :return: The database behavior of the column.
        """
        if self.db_name == "id":
            return text_dict["databaseBehaviorFilledByDatabase"]
        elif self.scenario_dependent:
            return text_dict["databaseBehaviorScenarioDependent"]
        else:
            return text_dict["databaseBehaviorNoBehavior"]

    @classmethod
    def dict_from_xml(cls, columns: xml.etree.ElementTree.Element, cell_style_dict: dict[str, CellStyle]) -> dict[str, Self]:
        """
        Converts a list of XML column elements into a dictionary of Column objects.
        :param columns: XML element containing column definitions.
        :param cell_style_dict: Dictionary mapping cell style IDs to CellStyle objects.
        :return: A dictionary mapping column IDs to Column objects.
        """

        return_dict = {}

        try:
            for column in columns:
                column_id = column.get("id")
                readable_name = column.find("ReadableName").text
                description = column.find("Description").text
                unit = column.find("Unit").text
                column_width = float(column.find("ColumnWidth").text)
                cell_style = cell_style_dict[column.find("CellStyle").text] if column.find("CellStyle").text is not None else None

                return_dict[column_id] = Column(readable_name=readable_name,
                                                db_name=column_id,
                                                description=description,
                                                unit=unit,
                                                column_width=column_width,
                                                cell_style=cell_style)
        except KeyError as e:
            missing_styles = []
            for column in columns:
                if column.find("CellStyle").text not in cell_style_dict and column.find("CellStyle").text is not None:
                    if column.find("CellStyle").text not in missing_styles:
                        missing_styles.append(column.find("CellStyle").text)

            raise ValueError(f"Cell style definition(s) {missing_styles} not found for column '{column_id}' in the xml-file. Please define it/them.")

        return return_dict


class ExcelDefinition:
    def __init__(self, file_name: str, version: str, sheet_header: str, description_row_height: float, columns: list[Column]):
        """
        Represents a configuration for a spreadsheet.

        :param file_name: The name of the file.
        :param version: The version identifier of the spreadsheet configuration.
        :param sheet_header: The header of the spreadsheet.
        :param description_row_height: Height of the description row in the spreadsheet.
        :param columns: List of columns of the spreadsheet.
        """

        self.file_name = file_name
        self.version = version
        self.sheet_header = sheet_header
        self.columns = columns
        self.description_row_height = description_row_height

    @classmethod
    def dict_from_xml(cls, excel_definitions: xml.etree.ElementTree.Element, column_dict: dict[str, Column], color_dict: dict[str, Color]) -> dict[str, Self]:
        """
        Converts a list of XML excel definition elements into a dictionary of ExcelDefinition objects.
        :param excel_definitions: XML element containing excel definition definitions.
        :param column_dict: Dictionary mapping column IDs to Column objects.
        :param color_dict: Dictionary mapping color IDs to Color objects.
        :return: A dictionary mapping excel definition IDs to ExcelDefinition objects.
        """
        return_dict = {}

        for excel_definition in excel_definitions:
            file_name = excel_definition.get("id")
            version = excel_definition.find("Version").text
            sheet_header = excel_definition.find("SheetHeader").text
            description_row_height = float(excel_definition.find("DescriptionRowHeight").text)
            columns = []

            try:
                for column in excel_definition.find("Columns"):
                    column_definition = column_dict[column.get("id")]
                    columns.append(column_definition.get_copy_with_scenario_dependent(column.get("scenarioDependent") == "True", color_dict))
            except KeyError as e:
                missing_columns = []
                for column in excel_definition.find("Columns"):
                    if column.get("id") not in column_dict:
                        missing_columns.append(column.get("id"))

                raise ValueError(f"Column definition(s) {missing_columns} not found for excel definition '{file_name}' in the xml-file. Please define it/them.")

            return_dict[file_name] = ExcelDefinition(file_name=file_name,
                                                     version=version,
                                                     sheet_header=sheet_header,
                                                     description_row_height=description_row_height,
                                                     columns=columns)

        return return_dict
